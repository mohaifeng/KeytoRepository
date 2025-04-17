import openpyxl
import re


def ColCount(n):
    """
    将第n列转为excel格式列数：A->AA->AAA...
    :param n: 第n列
    :return: str：第n列对应字母值
    """
    result = ''
    while n > 0:
        n -= 1
        result = chr(n % 26 + 65) + result
        n //= 26
    return result


class EXCEL:

    def __init__(self, file_path, excel_name):
        """
        :param file_path: 文件路径
        :param excel_name: 文件名称
        """
        self.cell_value = ''
        self.row = 0
        self.col = 0
        self.sheet_names = []
        self.test_sheet = ''
        self.dim = None
        self.file_path = file_path
        self.excel_name = excel_name
        self.excel_sheet = None
        # os.chdir(self.file_path)  # 修改工作路径
        self.work_book = openpyxl.load_workbook(self.excel_name)  # 返回一个workbook数据类型的值

    def Read_Sheets(self):
        """
        读取excel所有sheet
        :return: 单元格内容字符串
        """
        self.sheet_names = self.work_book.sheetnames  # 打印Excel表中的所有sheet
        return self.sheet_names

    def SHEET_ATTRIBUTE(self, sheet_name):
        """
        获取指定sheet相关属性
        :param sheet_name: sheet name
        """
        self.excel_sheet = self.work_book[sheet_name]  # 获取指定sheet表
        # print(self.excel_sheet, type(self.excel_sheet))
        self.row = self.excel_sheet.max_row  # 获取行数
        # print(self.row,type(self.row))
        self.col = self.excel_sheet.max_column  # 获取列数
        # print(self.col,type(self.col))
        self.dim = self.excel_sheet.dimensions  # 获取表格的尺寸大小
        # print(self.dim,type(self.dim))

    def READ_CELL(self, form):
        """
        读取指定单元格内容
        :param form: 指定单元格
        :return: 单元格内容-->str
        """
        # self.test_sheet = input("")
        self.cell_value = self.excel_sheet[form].value  # 获取单元格的数据
        # print(cell.value)
        # cell2 = sheet.cell(row=1, column=1)  # 获取第1行第1列的数据，和上面一致
        # print(cell1.value, type(cell1.value))  # 调试
        return self.cell_value

    def READ_CELL_LINES(self):
        """
        将单元格内多行内容按行输出为字典形式
        :return: 返回多行内容-->字典
        """
        line_lst = []
        step_line = self.cell_value.splitlines(False)
        for single_line in step_line:
            if 'Yes' not in single_line:
                line_number = single_line.split('.')[0]
                line_lst.append(int(line_number))
            else:
                single_index = step_line.index(single_line)
                step_line[single_index - 1] += step_line[single_index]
                del step_line[single_index]
        line_dic = dict(zip(line_lst, step_line))
        return line_dic

    # 数据处理，依次提取单元格测试步骤指令字符串，并输出列表
    def EXTRACT_SEND_COMMAND(self):
        """
        按行提取单元格测试步骤指令字符串，并输出字典
        :return: 指令->字典
        """
        num = 0
        lst_send_command = {}
        for line in self.cell_value.splitlines():
            num += 1
            lst = re.findall(r"[(](.*?)[)]", line)
            if lst:
                lst_send_command[num] = lst

        return lst_send_command


if __name__ == '__main__':
    print(bool('0'))
