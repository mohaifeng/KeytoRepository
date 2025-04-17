from binascii import a2b_hex
import sys
import datetime
import time
import openpyxl
import Com.Port.serialport as sp
from Com.Port.check import CRC_16
import numpy as np


def MadpCmdConfig(operat_order: str, madp_command: str):
    """
    MADP脚本指令转化为oem指令
    :param operat_order: 操作指令
    :param madp_command: 指令字符串
    :return: OEM协议格式指令字符串
    """
    data_hex = madp_command.encode().hex().replace('0x', '')
    operat_order_hex = operat_order.encode().hex().replace('0x', '').zfill(2)
    data_len_hex = hex(len(data_hex) // 2).replace('0x', '').zfill(4)
    oem_data = 'AA' + operat_order_hex + data_len_hex + data_hex
    crc16 = CRC_16(oem_data)
    fin_oem_data = oem_data + crc16
    return fin_oem_data


def MadpSend(order: str, string: str):
    """
    :param order: 操作指令
    :param string: 指令字符串
    :return: 0:返回数据正确 1:状态错误
    """
    cmd_str = MadpCmdConfig(order, string)
    sp.ser.PortSend(bytes.fromhex(cmd_str))
    # print(datetime.datetime.now(), end=':')
    # print('Send:', cmd_str, '->', string)


def MadpReceive(timeout=2000):
    rec_data = ''
    count = 0
    while True:
        rec_dat = sp.ser.PortReceive_word(timeout)
        if rec_dat:
            rec_dat = rec_dat.hex().upper()
        crc_data = CRC_16(rec_data)
        if crc_data == rec_dat:
            rec_data += rec_dat
            print(datetime.datetime.now(), end=':')
            print('Receive:', rec_data)
            return rec_data
        elif rec_dat == 0:
            if count == 0:
                print(datetime.datetime.now(), end=':')
                print('数据无返回！')
                return 0
            else:
                print(datetime.datetime.now(), end=':')
                print('数据返回错误:', rec_data)
                return 1
        else:
            count += 1
            rec_data += rec_dat
            continue


def Sys_State(sta):
    stat_dict = {0: '空闲/无错误', 1: '执行成功', 2: '运行中', 3: '暂停', 4: '执行完成', 10: '运行中禁止操作',
                 11: '操作指令执行失败', 12: '操作指令无效', 15: '读寄存器错误', 16: '写寄存器错误',
                 17: '读对象字典错误', 18: '写对象字典错误', 20: '流程指令无效',
                 21: '流程指令语法错误', 22: '流程指令执行超时', 23: '流程指令执行错误', 30: '节点报警'}
    if sta in stat_dict.keys():
        print(stat_dict[sta])
    else:
        print('无错误状态信息！')


def QaCheckState():
    q = 0
    check_command = "AA710000E771"
    print('检查状态！')
    while True:
        sp.ser.PortSend(bytes.fromhex(check_command))
        print(datetime.datetime.now(), end=':')
        print('Send:', check_command)
        get_data = MadpReceive()
        print(datetime.datetime.now(), end=':')
        if get_data == 0:
            if q < 1:
                print('无回复数据，重发指令！')
                q += 1
                continue
            else:
                print('重发失败！无回复数据')
                break
        else:
            re = get_data
            print('Receive:', re)
            sys_stat = int(re[4:6], 16)
            if sys_stat == 2:
                time.sleep(0.01)
                continue
            elif sys_stat == 0:
                break
            else:
                print('状态错误:', end='')
                Sys_State(sys_stat)


# def Get_pressure(address: int):
#     can_pressure = str(address) + ',2000.4'
#     com = bytes.fromhex(MadpCmdConfig('c', can_pressure))
#     ser.PortSend(com)
#     rec = ser.PortReceive()
#     if rec:
#         num = int(a2b_hex(rec.hex()[10:-4]).decode().split(',')[1])
#     else:
#         print('无回复气压数据')
#         return 0
#     return num


# def Get_pressure_slope(address: int):
#     """
#     读取ADP气压斜率
#     :param address: ADP地址
#     :return: 气压斜率：int
#     """
#     pressureslope_cmd = str(address) + ',2000.17'
#     com = bytes.fromhex(MadpCmdConfig('c', pressureslope_cmd))
#     ser.PortSend(com)
#     rec = ser.PortReceive()
#     if rec:
#         num = int(a2b_hex(rec.hex()[10:-4]).decode().split(',')[1])
#     else:
#         print('无回复气压斜率数据')
#         return 0
#     return num


def Get_Zaxle_Height(address: int):
    zaxleheight_cmd = str(address) + ',2000.101'
    com = bytes.fromhex(MadpCmdConfig('c', zaxleheight_cmd))
    sp.ser.PortSend(com)
    rec = sp.ser.PortReceive_byte()
    if rec:
        num = int(a2b_hex(rec.hex()[10:-4]).decode().split(',')[1])
    else:
        print('无回复Z轴高度数据')
        return 0
    return num


def Creatwb(wbname_):
    wb = openpyxl.Workbook()
    wb.save(filename=wbname_)
    print('新建Excel:' + wbname_ + '成功')


def SaveToExcel(data, wbname_, sheet_name, r_, col_):
    print('写入excel:')
    wb = openpyxl.load_workbook(filename=wbname_)
    sheet = wb.active
    sheet.title = sheet_name
    sheet.cell(row=r_, column=col_, value=str(data))
    wb.save(filename=wbname_)
    print('保存成功')


if __name__ == '__main__':
    # ser = sp.SERIALPORT('com16', 38400)
    # fig = plt.figure(1)
    # cmd_list=['1Az500,100,0']
    cmd_list = '0Sp9000,5000|1-8Aw43,1|41-48Zp130000,150000|41*Zg50000,100|0L200|42*Zg50000,100|0L200|43*Zg50000,' \
               '100|0L200|44*Zg50000,100|0L200|45*Zg50000,100|41*Zp130000,50000|0L200|46*Zg50000,100|42*Zp130000,' \
               '50000|0L200|47*Zg50000,100|43*Zp130000,50000|0L200|48*Zg50000,100|44*Zp130000,' \
               '50000|0L200|45*Zp130000,50000|0L200|46*Zp130000,50000|0L200|47*Zp130000,50000|0L200|48*Zp130000,' \
               '50000|0L200|41-48X5000|41-48Zp0,150000'
    # num_count = 0
    # for k in range(0, 10):
    #     numm = Get_pressure(1)
    #     num_count += numm
    # num_count = num_count // 10
    # num_count1 = num_count
    # print('num_count', num_count)

    # MadpSend('E', '1*Ai1000,1,0')
    MadpSend('E', cmd_list)
    # while True:
    #     check_slope1 = Get_pressure_slope(1)
    #     print(Get_pressure(1))
    #     print("check_slope1:", check_slope1)
    #     if check_slope1 > 300:
    #         if Get_pressure_slope(1) > 300:
    #             print('探测到液面')
    #             break
    # QaCheckState()
    # MadpSend('E', '1*Ai1000,1,0')
    # MadpSend('E', '41*Zp136000,20000')
    # check_slope2_list = []
    # check_pressure_list = []
    # i = 0
    # cmd_list = ['1Ai100,1000', '1Ae100,0,1000']
    # while True:
    #     for cmd in cmd_list:
    #         MadpSend('E', cmd)
    #     check_pressure1 = Get_pressure(1)
    #     for cmd in cmd_list:
    #         MadpSend('E', cmd)
    #     check_pressure2 = Get_pressure(1)
    #     print("check_pressure1:", check_pressure1)
    #     print("check_pressure2:", check_pressure2)
    #     pre_difference = check_pressure2 - check_pressure1
    #     if check_pressure2 > 30000:
    #         if pre_difference < 2000:
    #             i += 1
    #             if i > 2:
    #                 print('探测成功')
    #                 MadpSend('E', '41Zt0')
    #                 break
    # QaCheckState()
    # MadpSend('E', '41Zz50000')
    # QaCheckState()
    # MadpSend('E', '1Az500,100,2')
# print("check_slope2:", check_slope2)
# check_pressure = Get_pressure(1)
# print("check_pressure:", check_pressure)
# # check_slope2_list.append(check_slope2)
# # check_pressure_list.append(check_pressure)
# if check_slope2 > 100:
#     if check_pressure > 4000:
#         MadpSend('E', '41Zt0')
#         print('height:', Get_Zaxle_Height(41))
#         MadpSend('E', '41Zp0,50000')
#         QaCheckState()
#         MadpSend('E', '1Az500,100,2')
#         break
# ax1 = plt.subplot()
# plt.plot(check_slope2_list)
# ax2 = plt.subplot()
# plt.plot(check_pressure_list)
# plt.show()
#     while True:
#         check_pressure1 = Get_pressure(1) - num_count1
#         check_pressure2 = Get_pressure(1) - num_count1
#         check_pressure3 = (check_pressure1 + check_pressure2) // 2
#         print(datetime.datetime.now(), end=':')
#         print("check_pressure1:", check_pressure1)
#         print(datetime.datetime.now(), end=':')
#         print("check_pressure2:", check_pressure2)
#         if check_pressure3 < -1000 and check_pressure2 - check_pressure1 < -100:
#             break
#         elif check_pressure3 > 4000:
#             num_count1 = 4000
#             continue
#         else:
#             continue
#     while True:
#         get_pre2 = Get_pressure(1) - num_count
#         check_pressure9 = Get_pressure(1) - num_count1
#         check_pressure10 = Get_pressure(1) - num_count1
#         print(datetime.datetime.now(), end=':')
#         print('get_pre2:', get_pre2)
#         print(datetime.datetime.now(), end=':')
#         print('check_pressure9:', check_pressure9)
#         print(datetime.datetime.now(), end=':')
#         print('check_pressure10:', check_pressure10)
#         if check_pressure3 > get_pre2 > -7000 or check_pressure10 - check_pressure9 > 1000:
#             break
#         else:
#             continue
#     i = 0
#     while True:
#         get_pre3 = Get_pressure(1) - num_count
#         print(datetime.datetime.now(), end=':')
#         print('get_pre3:', get_pre3)
#         if get_pre3 > num_count:
#             if get_pre3 > 5000:
#                 get_pre4 = Get_pressure(1) - num_count
#                 get_pre5 = Get_pressure(1) - num_count
#                 print(datetime.datetime.now(), end=':')
#                 print('get_pre4:', get_pre4)
#                 print(datetime.datetime.now(), end=':')
#                 print('get_pre5:', get_pre5)
#                 get_pre6 = get_pre5 - get_pre4
#                 if 0 < get_pre6 < 500:
#                     i += 1
#                     if i > 2:
#                         print('探测成功')
#                         MadpSend('E', '41Zt0')
#                         QaCheckState()
#                         MadpSend('E', '1D4008,0,0')
#                         QaCheckState()
#                         break
#                 else:
#                     continue
#             else:
#                 continue
#         elif num_count > get_pre3 > -7000:
#             get_pre7 = Get_pressure(1) - num_count
#             get_pre8 = Get_pressure(1) - num_count
#             print(datetime.datetime.now(), end=':')
#             print('get_pre7:', get_pre7)
#             print(datetime.datetime.now(), end=':')
#             print('get_pre8:', get_pre8)
#             get_pre9 = get_pre8 - get_pre7
#             if 0 < get_pre9 < 1000:
#                 i += 1
#                 if i > 1:
#                     print('探测成功')
#                     MadpSend('E', '41Zt0')
#                     QaCheckState()
#                     MadpSend('E', '1D4008,0,0')
#                     QaCheckState()
#                     break
#             else:
#                 continue
#
# time.sleep(2)
# MadpSend('E', '1D4009,0,50')
# MadpSend('E', '41Zp0,50000')
